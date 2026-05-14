from io import BytesIO

from openpyxl import load_workbook

from app.rag.extractors.base_extractor import BaseExtractor
from app.rag.models.raw_document import RawDocument


class ExcelExtractor(BaseExtractor):
    """
    Excel extractor.

    Supports:
    - XLSX byte-stream ingestion
    - local filesystem fallback

    Future:
    - XLS support
    - object-storage streaming
    """

    def extract_text(
        self,
        raw_document: RawDocument,
    ) -> str:

        if not raw_document:
            raise ValueError("raw_document is required")

        file_type = (raw_document.file_type or "").lower()

        if file_type == "xls":
            raise ValueError(
                "XLS extraction is not implemented yet. "
                "Please upload XLSX for now."
            )

        if file_type != "xlsx":
            raise ValueError(
                f"Unsupported Excel file type: {raw_document.file_type}"
            )

        try:

            workbook = self._load_workbook(raw_document)

            text_parts = []

            for sheet in workbook.worksheets:

                text_parts.append(
                    f"\n--- Sheet: {sheet.title} ---"
                )

                for row in sheet.iter_rows(values_only=True):

                    row_values = []

                    for cell in row:

                        if cell is None:
                            continue

                        cell_text = str(cell).strip()

                        if cell_text:
                            row_values.append(cell_text)

                    if row_values:
                        text_parts.append(
                            " | ".join(row_values)
                        )

            workbook.close()

            extracted_text = "\n".join(text_parts).strip()

            if not extracted_text:
                raise ValueError(
                    f"No text extracted from Excel file: "
                    f"{raw_document.file_name}"
                )

            return extracted_text

        except Exception as ex:

            raise ValueError(
                f"Failed to extract Excel file "
                f"{raw_document.file_name}. "
                f"error={str(ex)}"
            ) from ex

    def _load_workbook(
        self,
        raw_document: RawDocument,
    ):

        # Preferred distributed-safe extraction
        if raw_document.content:

            return load_workbook(
                filename=BytesIO(raw_document.content),
                read_only=True,
                data_only=True,
            )

        # Local dev fallback
        if raw_document.local_path:

            return load_workbook(
                filename=raw_document.local_path,
                read_only=True,
                data_only=True,
            )

        # Future: storage_uri stream extraction
        if raw_document.storage_uri:

            raise NotImplementedError(
                "storage_uri extraction not implemented yet"
            )

        raise ValueError(
            f"No content available for Excel file: "
            f"{raw_document.file_name}"
        )